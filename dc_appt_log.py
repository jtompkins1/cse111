import tkinter as tk
from tkinter import ttk, Frame, Label, Button, messagebox
import openpyxl
import os

def main():
    # Change the filepath inside the quotations below to match the location where you will save
    # the new dc.xlsx file that will be created.
    # Ensure this will be in the same folder as this program file dc_appt_log.py.
    path = (r"C:\Users\Jennifer\Documents\jen_school\byui\programming\cse111_programing_with_functions\cse111\dc.xlsx")

    root = tk.Tk()
    frame = Frame(root)
    frame.master.title("DC Scheduling Log")
    frame.pack()

    populate_main_window(frame, path)

    root.mainloop()

def create_dc_file(path):
    """Create a new xlss file with column headings "DC and "Delivery
    and save file
    Parameter path
    Return path"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    heading = ["DC", "Delivery#"]
    sheet.append(heading)
    path = workbook.save(path)

    return path

def populate_main_window(frame, path):
    """Populate the main window of this program. In other words, put
    the labels, text entry boxes, and buttons into the main window.

    Parameters
        frame: the main frame (window)
        path: the filepath where the dc.xlsx file is located
    Return: nothing
    """
    create_dc_file(path)

    widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
    widgets_frame.grid(row=0, column=0, padx=20, pady=10)

    dc_entry = ttk.Entry(widgets_frame)
    dc_entry.insert(0,"DC")
    # when you click on this entry it clears the text
    dc_entry.bind("<FocusIn>", lambda e: dc_entry.delete("0", "end"))
    dc_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

    # drop down selection
    dc_list = ["Select DC#", 6003, 6018, 7035, 7034, 7026, 6094, 6017, 6011, 6020, 6030, 6040]
    dc_entry = ttk.Combobox(widgets_frame, values=dc_list)
    dc_entry.current(0)
    dc_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

    del_entry = ttk.Entry(widgets_frame)
    del_entry.insert(0,"Delivery#")
    # when you click on this entry it clears the text
    del_entry.bind("<FocusIn>", lambda e: del_entry.delete("0", "end"))
    del_entry.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

    def load_data():
        """gets the data from excel file and loads it into the treeview
        parameters: none
        return: nothing"""
        
        if os.path.exists(path):
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active

            list_values = list(sheet.values)
            
            # add column names to treeview
            for col_name in list_values[0]:
                treeview.heading(col_name, text=col_name)
            #skip row one
            for value_tuple in list_values[1:]:
                treeview.insert("", tk.END, values=value_tuple)

        else:
            ttk.tkinter.messagebox.showwarning(title="Error", 
                                               message="File not found. Check location of file dc.xlsx and filepath entered on line 10 of dc_appt_log.py")

    def insert_row():
        """Extracts the values from the dc_entry and del_entry widgets and places
        them in a list, loads the existing excel file and appends the list into
        the last row of the worksheet, then saves worksheet and inserts the new
        row into the treeview widget. Once enterered the users input is cleared
        from the entry widgets.  
        parameters: none
        return: nothing"""

        # Extract data entered in dc_entry and del_entry widgets
        dc = dc_entry.get()
        delnum = del_entry.get()

        # insert row into excel sheet
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_values = [dc, delnum]
        sheet.append(row_values)
        workbook.save(path)
        
        # insert row into treeview
        treeview.insert("", tk.END, values=row_values)

        # clear values
        dc_entry.set(dc_list[0])
        del_entry.delete(0, "end")
        del_entry.insert(0, "Delivery#")

    # insert button
    button = ttk.Button(widgets_frame, text="Insert", command=insert_row)
    button.grid(row=2,  column=0, padx=5, pady=(0, 5), sticky="news")

    # create treeview frame
    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)

    # create treeview scrollbar
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")

    # create columns
    cols = ("DC", "Delivery#")

    # create treeview
    treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=20)

    # set width of treeview columns
    treeview.column("DC", width=100)
    treeview.column("Delivery#", width=100)

    # Configure column headings
    treeview.heading("#0", text="Item")
    treeview.heading("DC", text="DC")
    treeview.heading("Delivery#", text="Delivery#")

    # position treeview
    treeview.pack()
    # config scrollbar
    treeScroll.config(command=treeview.yview)

    # load data into treeview
    load_data()
    

if __name__ == "__main__":
    main()